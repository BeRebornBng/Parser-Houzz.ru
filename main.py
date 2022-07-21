import os
import time

import requests
import json
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
import configparser
from concurrent.futures import ThreadPoolExecutor

url = "https://www.houzz.ru/professionals/"
spec = ''
index = ''

mass = []

sh1: vars()
wb: vars()
links_wb: vars()
linksSh: vars()

def Func(value):
    value = "Hello my name is Bakit."
    return value

def Parser(indexLink):
    global sh1
    link = linksSh[indexLink][0].value
    response = requests.get(link)
    soup = BeautifulSoup(response.text, 'html.parser')

    tel = json.loads(soup.find('div', 'hz-page-content-wrapper').find('script').text)

    myDict = {'Наименование': '',
              'Сайт HOUZ': '',
              'Сайт': '',
              'Номер телефона': '',
              'Адрес': '',
              'О нас': '',
              'Предоставляемые услуги': '',
              'География работ': '',
              'Средняя стоимость работ': ''
              }
    try:
        myDict['Наименование'] = (tel[0]['name'])
    except:
        pass
    try:
        myDict['Сайт HOUZ'] = link
        myDict['Сайт'] = (tel[0]['url'])
    except:
        pass
    try:
        myDict['Номер телефона'] = (tel[0]['telephone'])
    except:
        pass
    try:
        myDict['Адрес'] = (tel[0]['address']['addressLocality'] + ' ' + tel[0]['address']['postalCode'])
    except:
        try:
            avgCost = soup.findAll('div', 'sc-183mtny-0 Row___StyledBox-sc-1xmq8tu-0 jqaQXH kwUbop')
            myDict['Адрес'] = avgCost[1].text
        except:
            pass
        pass
    try:
        myDict['О нас'] = (tel[0]['description'])
    except:
        pass
    try:
        myDict['Предоставляемые услуги'] = (tel[0]['hasOfferCatalog']['name'])
    except:
        pass
    try:
        myDict['География работ'] = (tel[0]['areaServed']['name'])
    except:
        pass
    try:
        myAvg = soup.findAll('div', 'sc-183mtny-0 Row___StyledBox-sc-1xmq8tu-0 jqaQXH kwUbop')
        myDict['Средняя стоимость работ'] = myAvg[2].text
    except:
        pass
    sh1.append((myDict['Наименование'], myDict['Сайт HOUZ'], myDict['Сайт'],
                myDict['Номер телефона'], myDict['Адрес'], myDict['О нас'],
                myDict['Предоставляемые услуги'], myDict['География работ'],
                myDict['Средняя стоимость работ']))
    print("Parser")
    print(indexLink)


count = 1


def getLinks(value):
    global linksSh
    global count
    global mass
    config = configparser.ConfigParser()
    config.read('MyData.ini')
    spec = config['Link']['spec']
    index = '/c/' + config['Link']['index']
    if (index == '/c/'):
        response = requests.get(url + spec + '/p/' + str(value))
    else:
        response = requests.get(url + spec + index + '/p/' + str(value))
    soup = BeautifulSoup(response.text, 'html.parser')
    s = soup.find('ul', 'hz-pro-search-results').findAll('a', 'hz-pro-ctl')
    for item in s:
        print("GetLink")
        print(count)
        linksSh.append([str(item.get('href'))])
        mass.append(count)
        count += 1


def getNumberOfPages():
    global spec
    global index
    config = configparser.ConfigParser()
    config.read('MyData.ini')
    spec = config['Link']['spec']
    index = '/c/' + config['Link']['index']
    print(url + spec + index)
    response = requests.get(url + spec + index)
    soup = BeautifulSoup(response.text, 'html.parser')
    s = soup.find('div', 'hz-pro-search-controls__pagination mlm').findAll('span', 'text-bold')
    numAll = ''
    numOne = ''
    numZero = ''
    for symbol in s[0].text:
        if (symbol != "\xa0" and symbol != ' '):
            numZero += symbol
    for symbol in s[1].text:
        if (symbol != "\xa0" and symbol != ' '):
            numOne += symbol
    for symbol in s[2].text:
        if (symbol != "\xa0" and symbol != ' '):
            numAll += symbol
    print(numAll + ' ' + numOne + ' ' + numZero)
    return (int(numAll) / ((int(numOne) - int(numZero)) + 1))


def start():
    global wb
    global sh1
    global links_wb
    global linksSh
    global count

    try:
        os.remove('output.xlsx')
        wb = Workbook()
        wb['Sheet'].title = "List"
        sh1 = wb.active
    except:
        links_wb = Workbook()
        links_wb['Sheet'].title = "List"
        linksSh = links_wb.active
        wb = Workbook()
        wb['Sheet'].title = "List"
        sh1 = wb.active

    try:
        os.remove('Links.xlsx')
        links_wb = Workbook()
        links_wb['Sheet'].title = "List"
        linksSh = links_wb.active
    except:
        links_wb = Workbook()
        links_wb['Sheet'].title = "List"
        linksSh = links_wb.active

    config = configparser.ConfigParser()
    config.read('MyData.ini')
    threads = int(config['Link']['threads'])

    try:
        numberOfPages = getNumberOfPages()
    except Exception:
        print("Ничего не найдено")
        return 1

    arr = []
    i = 0
    while (i < numberOfPages * 15):
        arr.append(i)
        i += 15

    with ThreadPoolExecutor(max_workers=(os.cpu_count() * threads)) as p:
        p.map(getLinks, arr)
    links_wb.save("Links.xlsx")
    time.sleep(5)

    links_wb = openpyxl.open("Links.xlsx", read_only=True)
    linksSh = links_wb.active
    sh1.append(('Наименование', 'Сайт HOUZ', 'Сайт',
                'Номер телефона', 'Адрес', 'О нас',
                'Предоставляемые услуги', 'География работ',
                'Средняя стоимость работ'))
    sh1[1][0].font = Font(bold=True)
    sh1[1][1].font = Font(bold=True)
    sh1[1][2].font = Font(bold=True)
    sh1[1][3].font = Font(bold=True)
    sh1[1][4].font = Font(bold=True)
    sh1[1][5].font = Font(bold=True)
    sh1[1][6].font = Font(bold=True)
    sh1[1][7].font = Font(bold=True)
    sh1[1][8].font = Font(bold=True)


    with ThreadPoolExecutor(max_workers=(os.cpu_count() * threads)) as p:
        p.map(Parser, mass)

    wb.save('output.xlsx')
    mass.clear()
    count = 1
    return 0